"""
Data Parser v3 – Moodys Orbis 4-Sheet Excel + Generic Excel
─────────────────────────────────────────────────────────────────────────────
Extracts full 10-year historical time series from Moodys Orbis exports.
Builds HistoricalYear objects for use in HistoricalAnalyzer.
─────────────────────────────────────────────────────────────────────────────
"""

import pandas as pd
import numpy as np
import sys, os
from datetime import date
from typing import Optional

_ROOT = os.path.dirname(os.path.abspath(__file__))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from finance_engine import (
    CompanyInputs, HistoricalYear, HistoricalMetrics, HistoricalAnalyzer
)


# ─────────────────────────────────────────────
# DISPLAY LABELS
# ─────────────────────────────────────────────

REQUIRED_FIELDS = {
    "revenue":              "Umsatz / Revenue",
    "ebitda":               "EBITDA",
    "ebit":                 "EBIT",
    "depreciation":         "D&A / Abschreibungen",
    "interest_expense":     "Zinsaufwand / Interest Expense",
    "tax_rate":             "Steuerquote / Tax Rate",
    "net_income":           "Jahresüberschuss / Net Income",
    "total_debt":           "Gesamtschulden / Total Debt",
    "cash":                 "Cash / Zahlungsmittel",
    "net_working_capital":  "Net Working Capital",
    "capex":                "CapEx / Investitionen",
}


# ─────────────────────────────────────────────
# ROW LABEL ALIASES  (DE + EN, Moodys Orbis)
# ─────────────────────────────────────────────

ROW_ALIASES = {
    "revenue": [
        "betriebsertrag (umsatz)",
        "umsatz",
        "umsatzerlöse",
        "net sales",
        "revenue",
        "total revenue",
        "net revenue",
        "sales",
        "operating revenue",
    ],
    "ebitda": [
        "ebitda",
        "earnings before interest taxes depreciation amortization",
    ],
    "ebit": [
        "betriebsgewinn/-verlust [=ebit]",
        "ebit",
        "operating profit",
        "operating income",
        "operating result",
        "betriebsergebnis",
    ],
    "depreciation": [
        "wertminderungen & abschreibungen",
        "abschreibungen",
        "afa",
        "d&a",
        "depreciation",
        "depreciation & amortization",
        "amortization",
    ],
    "interest_expense": [
        "zinsaufwand",
        "finanzaufwendungen",
        "interest expense",
        "finance costs",
        "net interest expense",
    ],
    "tax_rate": [
        "∟ steuern",
        "tax expense",
        "income taxes",
        "steueraufwand",
    ],
    "pretax_profit": [
        "gewinn/verlust vor steuern",
        "profit before tax",
        "income before tax",
        "ebt",
    ],
    "net_income": [
        "jahresüberschuss/-fehlbetrag",
        "gewinn/verlust nach steuern",
        "net income",
        "net profit",
        "profit after tax",
        "jahresüberschuss",
    ],
    "total_debt": [
        "langfristige finanzschulden",
        "kurzfristige finanzschulden",
        "vergebene kredite",
        "total debt",
        "financial debt",
        "gross debt",
        "gesamtschulden",
        "bankverbindlichkeiten",
    ],
    "cash": [
        "zahlungsmittel & zahlungsmitteläquivalente",
        "cash",
        "cash & equivalents",
        "kasse",
        "kassenbestand",
        "flüssige mittel",
    ],
    "net_working_capital": [
        "∟ working capital",
        "working capital",
        "net working capital",
        "nwc",
    ],
    "capex": [
        "capex",
        "capital expenditure",
        "capital expenditures",
        "sachanlagen-zugänge",
        "zugänge sachanlagen",
        "anlagezugänge",
        "additions to fixed assets",
        "investitionen in sachanlagen",
        "purchase of property",
    ],
}

MOODYS_SHEET_PRIORITY = {
    "revenue":           ["GuV-Rechnung", "P&L", "Income Statement"],
    "ebitda":            ["GuV-Rechnung", "P&L", "Income Statement"],
    "ebit":              ["GuV-Rechnung", "P&L", "Income Statement"],
    "depreciation":      ["GuV-Rechnung", "P&L", "Income Statement"],
    "interest_expense":  ["GuV-Rechnung", "P&L", "Income Statement"],
    "tax_rate":          ["GuV-Rechnung", "P&L", "Income Statement"],
    "pretax_profit":     ["GuV-Rechnung", "P&L", "Income Statement"],
    "net_income":        ["GuV-Rechnung", "P&L", "Income Statement"],
    "total_debt":        ["Bilanz", "Balance Sheet"],
    "cash":              ["Bilanz", "Balance Sheet"],
    "net_working_capital": ["Bilanz", "Balance Sheet"],
    "capex":             ["GuV-Rechnung", "P&L", "Bilanz", "Balance Sheet"],
}


# ─────────────────────────────────────────────
# CURRENCY / UNIT DETECTION
# ─────────────────────────────────────────────

def detect_currency_and_unit(unit_string: str) -> dict:
    s = str(unit_string).lower().strip()
    currency = "USD" if "usd" in s else "EUR" if "eur" in s else "?"
    if any(x in s for x in ["tsd", "tausend", "thousand"]):
        unit_label, scale_to_millions, raw_unit = "Tausend", 0.001, "tsd"
    elif any(x in s for x in ["mio", "million"]):
        unit_label, scale_to_millions, raw_unit = "Millionen", 1.0, "mio"
    elif any(x in s for x in ["mrd", "milliard", "billion"]):
        unit_label, scale_to_millions, raw_unit = "Milliarden", 1000.0, "mrd"
    else:
        unit_label, scale_to_millions, raw_unit = "Einzelwert", 0.000001, "units"
    return {
        "currency":         currency,
        "unit_label":       unit_label,
        "raw_unit":         raw_unit,
        "scale_to_millions": scale_to_millions,
        "display":          f"{currency} ({unit_label})",
        "symbol":           "€" if currency == "EUR" else "$" if currency == "USD" else "",
    }


def excel_serial_to_year(serial) -> Optional[int]:
    try:
        s = int(float(str(serial)))
        if 40000 < s < 50000:
            import datetime
            d = date(1899, 12, 30) + datetime.timedelta(days=s)
            return d.year
        if 2000 <= s <= 2035:
            return s
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────
# MOODYS ORBIS PARSER
# ─────────────────────────────────────────────

class MoodysOrbisParser:
    """
    Parses Bureau van Dijk / Moodys Orbis 4-sheet Excel export.
    Extracts full historical time series → builds HistoricalYear list.
    """

    def __init__(self):
        self.sheets:        dict = {}
        self.company_name:  str  = "Unknown"
        self.currency_info: dict = {}
        self.years:         list = []
        self.raw_data:      dict = {}   # field → {year: value}
        self.warnings:      list = []
        self.is_moodys_format: bool = False

    def load(self, file) -> bool:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            self.sheets = {
                sname: list(wb[sname].iter_rows(values_only=True))
                for sname in wb.sheetnames
            }
        except Exception as e:
            self.warnings.append(f"Ladefehler: {e}")
            return False

        moodys_keys = {
            "Bilanz", "GuV-Rechnung", "Kennzahlen", "Cover",
            "Balance Sheet", "P&L", "Income Statement", "Ratios"
        }
        self.is_moodys_format = len(moodys_keys & set(self.sheets.keys())) >= 2
        if not self.is_moodys_format:
            return False

        self._extract_company_name()
        self._extract_years_and_currency()
        self._extract_all_fields()
        return True

    # ── Internal extraction ───────────────────

    def _extract_company_name(self):
        for sheet_key in ["Cover"] + list(self.sheets.keys()):
            if sheet_key in self.sheets:
                for row in self.sheets[sheet_key][:3]:
                    val = row[0] if row else None
                    if val and isinstance(val, str) and len(val.strip()) > 3:
                        self.company_name = val.strip()
                        return

    def _extract_years_and_currency(self):
        for sheet_name, rows in self.sheets.items():
            if sheet_name == "Cover" or len(rows) < 7:
                continue
            date_row = rows[4] if len(rows) > 4 else []
            unit_row = rows[5] if len(rows) > 5 else []
            years = [yr for cell in date_row[1:] if (yr := excel_serial_to_year(cell))]
            if years:
                self.years = years
            for cell in unit_row[1:]:
                if cell and str(cell).strip():
                    self.currency_info = detect_currency_and_unit(str(cell))
                    return
        if not self.currency_info:
            self.currency_info = detect_currency_and_unit("EUR")
            self.warnings.append("⚠️ Währung nicht erkannt – EUR Default")

    def _extract_all_fields(self):
        for field, sheet_priority in MOODYS_SHEET_PRIORITY.items():
            found = False
            for sheet_name in sheet_priority:
                if sheet_name in self.sheets:
                    result = self._find_field_in_sheet(field, self.sheets[sheet_name])
                    if result is not None:
                        self.raw_data[field] = result
                        found = True
                        break
            if not found:
                for sheet_name, rows in self.sheets.items():
                    if sheet_name == "Cover":
                        continue
                    result = self._find_field_in_sheet(field, rows)
                    if result is not None:
                        self.raw_data[field] = result
                        found = True
                        break

    def _find_field_in_sheet(self, field: str, rows: list) -> Optional[dict]:
        aliases = ROW_ALIASES.get(field, [])
        for row in rows:
            if not row or row[0] is None:
                continue
            label_clean = str(row[0]).strip().lower().lstrip(" \t\xa0∟").strip()
            for alias in aliases:
                if label_clean == alias or label_clean.endswith(" " + alias):
                    values = list(row[1:])
                    result = {}
                    for i, yr in enumerate(self.years):
                        if i < len(values):
                            v = self._safe_float(values[i])
                            if v is not None:
                                result[yr] = v
                    if result:
                        return result
        return None

    def _safe_float(self, val) -> Optional[float]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return None if (val != val) else float(val)
        s = str(val).strip().lower()
        if s in ("n.v.", "n.s.", "n/a", "", "-", "—"):
            return None
        try:
            return float(s.replace(".", "").replace(",", "."))
        except Exception:
            return None

    # ── Public API ────────────────────────────

    def get_timeseries_df(self) -> pd.DataFrame:
        """Full time series as DataFrame, newest year first"""
        if not self.years:
            return pd.DataFrame()
        data = {field: {yr: self.raw_data[field].get(yr) for yr in self.years}
                for field in self.raw_data}
        return pd.DataFrame(data, index=self.years).sort_index(ascending=False)

    def build_historical_years(self, selected_years: Optional[list] = None) -> list[HistoricalYear]:
        """
        Build HistoricalYear list from raw_data.
        selected_years: subset of self.years to use (None = all available)
        """
        use_years = sorted(selected_years or self.years)
        hist_years = []

        for yr in use_years:
            def get(field, default=0.0):
                d = self.raw_data.get(field, {})
                v = d.get(yr)
                return v if v is not None else default

            # Derive tax rate from expense / pretax
            tax_exp   = get("tax_rate")    # actually the expense amount
            pretax    = get("pretax_profit")
            if pretax and pretax > 0 and tax_exp > 0 and tax_exp < pretax:
                tax_rate = tax_exp / pretax
            else:
                tax_rate = 0.25

            # CapEx: use actual if found, else 75% of D&A (maintenance proxy)
            dep   = get("depreciation")
            capex = get("capex")
            if capex == 0.0:
                capex = dep * 0.75

            hist_years.append(HistoricalYear(
                year                = yr,
                revenue             = get("revenue"),
                ebitda              = get("ebitda"),
                ebit                = get("ebit"),
                depreciation        = dep,
                interest_expense    = get("interest_expense"),
                net_income          = get("net_income"),
                total_debt          = get("total_debt"),
                cash                = get("cash"),
                net_working_capital = get("net_working_capital"),
                capex               = capex,
                tax_rate            = tax_rate,
            ))

        return [y for y in hist_years if y.revenue > 0]

    def build_company_inputs(
        self,
        hist_metrics: HistoricalMetrics,
        latest_year_data: "HistoricalYear",
    ) -> tuple[CompanyInputs, list]:
        """
        Build CompanyInputs from normalized historical metrics.
        Uses latest year balance sheet + normalized P&L from historical average.
        """
        warnings = list(self.warnings)

        inputs = CompanyInputs(
            revenue             = hist_metrics.normalized_revenue,
            ebitda              = hist_metrics.normalized_ebitda,
            ebit                = (hist_metrics.normalized_ebitda
                                   - latest_year_data.depreciation),
            depreciation        = latest_year_data.depreciation,
            interest_expense    = latest_year_data.interest_expense,
            tax_rate            = latest_year_data.tax_rate,
            total_debt          = latest_year_data.total_debt,
            cash                = latest_year_data.cash,
            net_working_capital = latest_year_data.net_working_capital,
            capex               = hist_metrics.normalized_capex,
            company_name        = self.company_name,
            currency_display    = self.currency_info.get("display", ""),
            revenue_cagr_hist   = hist_metrics.revenue_cagr,
            ebitda_margin_avg   = hist_metrics.ebitda_margin_avg,
            capex_intensity     = hist_metrics.capex_intensity_avg,
            nwc_intensity       = hist_metrics.nwc_intensity_avg,
        )
        return inputs, warnings


# ─────────────────────────────────────────────
# GENERIC EXCEL PARSER (fallback)
# ─────────────────────────────────────────────

class GenericExcelParser:
    """Fallback for non-Moodys files. Single-year extraction."""

    def __init__(self):
        self.df:            Optional[pd.DataFrame] = None
        self.currency_info: dict = {}
        self.company_name:  str  = "Target"
        self.mapping:       dict = {}
        self.unmapped:      list = []
        self.warnings:      list = []

    def load(self, file) -> bool:
        try:
            xl = pd.ExcelFile(file, engine="openpyxl")
            preferred = [s for s in xl.sheet_names if any(
                kw in s.lower() for kw in ["guv", "p&l", "income", "bilanz", "balance"])]
            sheet = preferred[0] if preferred else xl.sheet_names[0]
            df = pd.read_excel(file, sheet_name=sheet, header=None, engine="openpyxl")
            self.df = self._set_header(df)
            self._detect_currency()
            return True
        except Exception as e:
            self.warnings.append(f"Fehler: {e}")
            return False

    def _set_header(self, df):
        for i, row in df.iterrows():
            vals = [str(v).lower() for v in row if pd.notna(v)]
            if any(kw in " ".join(vals) for kw in ["umsatz", "revenue", "ebitda"]):
                df.columns = df.iloc[i]
                return df.iloc[i+1:].reset_index(drop=True)
        df.columns = df.iloc[0]
        return df.iloc[1:].reset_index(drop=True)

    def _detect_currency(self):
        if self.df is None:
            self.currency_info = detect_currency_and_unit("EUR")
            return
        for col in self.df.columns:
            if any(x in str(col).lower() for x in ["usd", "eur", "tsd", "mio"]):
                self.currency_info = detect_currency_and_unit(str(col))
                return
        self.currency_info = detect_currency_and_unit("EUR")

    def auto_map(self) -> dict:
        if self.df is None:
            return {"mapped": {}, "unmapped": list(REQUIRED_FIELDS.keys()), "available_columns": []}
        available = list(self.df.columns)
        mapped, unmapped = {}, []
        for field, aliases in ROW_ALIASES.items():
            found = None
            for alias in aliases:
                for col in available:
                    if alias in str(col).lower():
                        found = col
                        break
                if found:
                    break
            if found:
                mapped[field] = found
            else:
                unmapped.append(field)
        self.mapping  = mapped
        self.unmapped = unmapped
        return {"mapped": mapped, "unmapped": unmapped, "available_columns": available}

    def build_company_inputs(self, manual_overrides={}) -> tuple:
        vals     = {**self._extract_latest(), **manual_overrides}
        warnings = list(self.warnings)

        def get(f, d=0.0):
            v = vals.get(f)
            if v is None:
                warnings.append(f"⚠️ '{f}' fehlt")
                return d
            return v

        ebitda = get("ebitda")
        dep    = get("depreciation")
        tax    = get("tax_rate", 0.25)
        if tax > 1:
            tax /= 100.0

        inputs = CompanyInputs(
            revenue             = get("revenue"),
            ebitda              = ebitda,
            ebit                = vals.get("ebit") or (ebitda - dep),
            depreciation        = dep,
            interest_expense    = get("interest_expense"),
            tax_rate            = tax,
            total_debt          = get("total_debt"),
            cash                = get("cash"),
            net_working_capital = get("net_working_capital"),
            capex               = get("capex") or dep * 0.75,
            company_name        = self.company_name,
            currency_display    = self.currency_info.get("display", ""),
        )
        return inputs, warnings, self.currency_info

    def _extract_latest(self) -> dict:
        if self.df is None or not self.mapping:
            return {}
        vals = {}
        for field, col in self.mapping.items():
            if col in self.df.columns:
                series = pd.to_numeric(self.df[col], errors="coerce").dropna()
                if not series.empty:
                    vals[field] = float(series.iloc[-1])
        return vals


# ─────────────────────────────────────────────
# UNIFIED ENTRY POINT
# ─────────────────────────────────────────────

def parse_file(file) -> tuple:
    """Auto-detects format. Returns (parser, is_moodys: bool)"""
    moodys = MoodysOrbisParser()
    if moodys.load(file):
        return moodys, True
    generic = GenericExcelParser()
    generic.load(file)
    return generic, False